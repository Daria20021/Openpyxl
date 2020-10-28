import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Таблица1"
ws['A1'] = 'ФИО'
ws['B1'] = 'Профвзносы'
ws['C1'] = 'Месяц'
ws['D1'] = 'Сумма'

ws.append(["имя1", "+", "оба", "20"])
ws.append(["имя2", "+", "оба", "20"])
ws.append(["имя3", "-", "сен", "10"])
ws.append(["имя4", "-", "окт", "13"])
ws.append(["имя5", "-", "сен", "0"])
ws.append(["имя6", "+", "оба", "10"])
ws.append(["имя7", "-", "окт", "20"])
ws.append(["имя8", "+", "оба", "20"])
ws.append(["имя9", "-", "окт", "19"])
ws.append(["имя10", "+", "оба", "15"])
ws.append(["имя11", "-", "ноя", "20"])
ws.append(["имя12", "-", "сен", "20"])

ws2 = wb.create_sheet("Таблица2")
ws2['A1'] = 'ФИО'
ws2['B1'] = 'Лагерь'
ws2['C1'] = 'Статус'

ws2.append(["имя1", "Яхта"])
ws2.append(["имя2", "Бухта"])
ws2.append(["имя3", "Яхта"])
ws2.append(["имя4", "Яхта"])
ws2.append(["имя5", "Бухта"])
ws2.append(["имя6", "Яхта"])
ws2.append(["имя7", "Бухта"])
ws2.append(["имя8", "Бухта"])
ws2.append(["имя9", "Яхта"])
ws2.append(["имя10", "Бухта"])
ws2.append(["имя11", "Яхта"])
ws2.append(["имя12", "Бухта"])

for i in range(2, 14):
    if ws.cell(row=i, column=2).value == "+" and ws.cell(row=i, column=3).value == "оба" and ws.cell(row=i,
                                                                                                     column=4).value == "20":
        ws2['С' + str(i)] = "+"
else:
    ws2['С' + str(i)] = "-"
wb.save("Отдых.xlsx")
